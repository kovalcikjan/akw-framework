"""
Faze 1A helper: vytvori prazdnou seeds_template.xlsx pro rucni sber seedu.

DULEZITE: Tento script NIKDY negeneruje keywords. Vytvari pouze strukturovany
XLSX template s popisem sloupcu a instrukcemi. User si sam dopise seedy z
Ahrefs, GSC, product feedu, Marketing Mineru atd.

Pouziti:
    python src/create_seeds_template.py --project-root /path/to/project
    # Vytvori: <project>/data/raw/seeds_template.xlsx

    python src/create_seeds_template.py --project-root . --client "Braun"
    # Pouzije client jmeno v nadpisech
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


COLUMNS = [
    ("keyword", "povinne", "Klicove slovo v puvodni forme (s diakritikou pokud relevantni)"),
    ("source", "povinne", "Zdroj: ahrefs / gsc / marketing_miner / competitor_<domena> / client_seed / manual / paa"),
    ("volume", "volitelne", "Mesicni search volume (cele cislo). Prazdne pokud nevis."),
    ("kd", "volitelne", "Keyword difficulty 0-100. Prazdne pokud nevis."),
    ("position", "volitelne", "Soucasna pozice klienta 1-100 (pokud rankuje). 21 = MM konvence 'nerankuje v top 20'."),
    ("url", "volitelne", "Ranking URL klienta (pokud rankuje). Prazdne pokud ne."),
    ("notes", "volitelne", "Libovolna poznamka (priorita, kontext, flag pro review)."),
]


HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True, size=11)


def build_seeds_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Hlavni sheet s hlavickami sloupcu — zadna data."""
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)

    for col_idx, (_, required, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = REQUIRED_FILL if required == "povinne" else HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    width_map = {"keyword": 40, "source": 25, "volume": 12, "kd": 8,
                 "position": 10, "url": 50, "notes": 40}
    for col_idx, (name, _, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_map.get(name, 18)

    ws.freeze_panes = "A2"


def build_instructions_sheet(ws: openpyxl.worksheet.worksheet.Worksheet,
                             client: str) -> None:
    """Instrukce pro uzivatele — jak sbirat seedy rucne."""
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 80

    title = f"AKW Seed Keywords Template — {client}"
    ws.append([title])
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.append([])

    ws.append(["Sloupce:"])
    ws.cell(row=ws.max_row, column=1).font = HEADER_FONT
    for name, required, desc in COLUMNS:
        ws.append([name, required, desc])
    ws.append([])

    ws.append(["Jak pouzit:"])
    ws.cell(row=ws.max_row, column=1).font = HEADER_FONT
    steps = [
        ("1.", "", "Otevri sheet 'seeds' a vkladej seedy rucne z techto zdroju:"),
        ("", "", "  - Ahrefs (klient + top 3-5 konkurenti: Top pages, Organic keywords)"),
        ("", "", "  - Google Search Console (existing queries klienta)"),
        ("", "", "  - Product feed / sitemap klienta"),
        ("", "", "  - Marketing Miner UI (expansion + volume)"),
        ("", "", "  - Vlastni brainstorm z briefu + business_research"),
        ("2.", "", "Do sloupce 'source' zapis odkud seed pochazi (pro traceability)."),
        ("3.", "", "Nemusis vyplnit volume/kd/position — doplni se pozdeji v Fazi 6.5."),
        ("4.", "", "Cilovy rozsah: 50-500 seedu (ne tisice — kvalita > kvantita)."),
        ("5.", "", "Az mas hotovo, uloz soubor (klidne pridej dalsi listy) a oznam AI."),
    ]
    for a, b, c in steps:
        ws.append([a, b, c])
    ws.append([])

    ws.append(["Co AI NESMI:"])
    ws.cell(row=ws.max_row, column=1).font = HEADER_FONT
    forbidden = [
        "  - Generovat seedy pres DFS MCP / Ahrefs MCP / jakykoliv keyword API",
        "  - Prepisovat / doplnovat radky v tomto souboru (data/raw je READ-ONLY pro AI)",
        "  - Spoustet scripty ktere automaticky plni seedy",
    ]
    for line in forbidden:
        ws.append(["", "", line])
    ws.append([])

    ws.append(["Co AI SMI:"])
    ws.cell(row=ws.max_row, column=1).font = HEADER_FONT
    allowed = [
        "  - V chatu navrhnout konkretni seedy na zaklade briefu + research",
        "  - Upozornit na mezery (napr. 'chybi ti seedy pro produkt X')",
        "  - Po uploadu data precist a reportovat distribuce zdroju",
    ]
    for line in allowed:
        ws.append(["", "", line])


def main() -> int:
    parser = argparse.ArgumentParser(description="Faze 1A: Vytvor prazdnou seeds_template.xlsx")
    parser.add_argument("--project-root", type=Path, required=True,
                        help="Cesta k projektu (obsahuje data/raw/)")
    parser.add_argument("--client", type=str, default="<CLIENT>",
                        help="Jmeno klienta pro nadpisy")
    parser.add_argument("--force", action="store_true",
                        help="Prepis existujici seeds_template.xlsx")
    args = parser.parse_args()

    project_root = args.project_root.resolve()
    target_dir = project_root / "data" / "raw"
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / "seeds_template.xlsx"

    if target_path.exists() and not args.force:
        print(f"EXISTUJE: {target_path}")
        print("Pouzij --force pokud chces prepsat (ztratis stavajici obsah).")
        return 1

    wb = openpyxl.Workbook()
    ws_seeds = wb.active
    ws_seeds.title = "seeds"
    build_seeds_sheet(ws_seeds)

    ws_instr = wb.create_sheet("instructions")
    build_instructions_sheet(ws_instr, args.client)

    wb.save(target_path)
    print(f"OK: {target_path}")
    print(f"Klient: {args.client}")
    print("Sheety: seeds (prazdny), instructions")
    return 0


if __name__ == "__main__":
    sys.exit(main())
